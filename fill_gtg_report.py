#!/usr/bin/env python3
"""
fill_gtg_report.py
==================
Remplit automatiquement un rapport Excel de releves de parametres GTG
(Gas Turbine Generator) a partir de photos d'ecrans HMI, en utilisant
l'API vision GRATUITE de Google Gemini pour lire les valeurs des capteurs.

FONCTIONNEMENT
--------------
1. Le script analyse la structure du fichier Excel et detecte automatiquement :
   - les capteurs avec un tag d'instrument (ex "TE-532-AB", "PT-537")
   - certains champs SANS tag d'instrument mais avec un nom de parametre
     (ex "Intensite generateur", "Vitesse turbine") -> voir SPECIAL_FIELDS
   ... et la cellule exacte ou chaque valeur doit etre ecrite, pour chacune
   des 6 heures de releve (4h, 8h, 12h, 16h, 20h, 24h).
2. Pour chaque photo d'ecran fournie, il envoie l'image a Gemini (vision)
   avec une consigne stricte : lire UNIQUEMENT les paires (tag, valeur)
   visibles a l'ecran, sans deviner ni inventer. Gere aussi :
   - le systeme "legende par lettre" (A,B,C... relie un tag a sa valeur
     situee ailleurs sur l'ecran)
   - les champs sans tag (parametres electriques generateur, vitesse
     turbine, moyenne echappement) via des pseudo-tags dedies
   - l'ecran "Overview" qui affiche plusieurs turbines a la fois : seule
     la turbine passee en --turbine est extraite
3. Chaque tag/pseudo-tag lu est mis en correspondance avec le classeur Excel
   (le prefixe de la turbine, ex "812-", est retire automatiquement des tags
   d'instrument classiques).
4. La valeur est ecrite dans la/les cellule(s) correspondant a l'heure de
   releve choisie (parametre --heure). Certains champs (intensite, tension)
   sont dupliques sur les 3 lignes de phase du classeur.
5. Un rapport texte est affiche listant :
      - les valeurs trouvees et ecrites avec succes
      - les valeurs lues sur les photos mais absentes du classeur
      - les capteurs attendus dans le classeur mais jamais vus sur une photo

Le script NE DEVINE JAMAIS une valeur. Si Gemini n'est pas sur de la valeur
lue, il ne la reporte pas et le capteur apparait dans la liste "non lus".

PREREQUIS (sur votre ordinateur)
---------------------------------
    pip install google-genai openpyxl

    Cle API GRATUITE (aucune carte bancaire requise) :
    1. Allez sur https://aistudio.google.com/apikey
    2. Connectez-vous avec un compte Google et cliquez "Create API key"
    3. Definissez-la comme variable d'environnement, ou passez --api-key

    Limites gratuites (largement suffisantes pour 18 photos par tournee) :
    ~10 requetes/minute avec gemini-3.5-flash. Le script attend et reessaie
    automatiquement si la limite est atteinte.

UTILISATION
-----------
    # 1) Generer la carte des capteurs (une seule fois par modele de rapport,
    #    a refaire si vous mettez a jour ce script)
    python fill_gtg_report.py map --excel GTG811.xlsx --out sensor_map.json

    # 2) Remplir un releve (ex: le releve de 12h de la turbine 812)
    python fill_gtg_report.py fill ^
        --excel GTG812.xlsx ^
        --map sensor_map.json ^
        --heure 12h ^
        --turbine 812 ^
        --images-dir Photos_812_12h ^
        --out GTG812_rempli.xlsx
"""

import argparse
import base64
import json
import os
import re
import sys
import time
from pathlib import Path

import openpyxl

try:
    from google import genai
    from google.genai import types as genai_types
except ImportError:
    genai = None
    genai_types = None

HEURES = ['4h', '8h', '12h', '16h', '20h', '24h']
TAG_PATTERN = re.compile(r'^[A-Z]{1,4}-\d{2,4}(?:-[A-Z0-9]{1,3})?$')


class QuotaExhaustedError(Exception):
    """Levee quand le quota GRATUIT quotidien semble reellement epuise
    (par opposition a une simple limite de debit temporaire, qui se resout
    avec une pause de quelques secondes)."""
    pass

# Champs qui n'ont PAS de tag d'instrument dans le classeur, juste un nom de
# parametre. On les detecte par le texte de leur libelle, et on leur donne un
# "pseudo-tag" fixe que Gemini doit utiliser tel quel dans sa reponse JSON.
#   label_contains  : sous-chaines (toutes requises, en MAJUSCULES) qui identifient le libelle
#   label_excludes  : sous-chaines qui, si presentes, excluent la cellule (evite les faux positifs)
#   expand_rows     : nombre de lignes consecutives (a partir du libelle) ou dupliquer
#                      la meme valeur - utilise pour Intensite/Tension (3 phases identiques)
SPECIAL_FIELDS = [
    {'pseudo_tag': 'GEN_INTENSITE',          'label_contains': ['INTENSITE'],                 'label_excludes': ['EXCITATION'], 'expand_rows': 3},
    {'pseudo_tag': 'GEN_TENSION',            'label_contains': ['TENSION'],                   'label_excludes': ['EXCITATION'], 'expand_rows': 3},
    {'pseudo_tag': 'GEN_FREQUENCE',          'label_contains': ['FREQUENCE'],                  'expand_rows': 1},
    {'pseudo_tag': 'GEN_PUISSANCE_ACTIVE',   'label_contains': ['PUISSANCE ACTIVE'],           'expand_rows': 1},
    {'pseudo_tag': 'GEN_PUISSANCE_REACTIVE', 'label_contains': ['PUISSANCE REACTIVE'],         'expand_rows': 1},
    {'pseudo_tag': 'GEN_FACTEUR_PUISSANCE',  'label_contains': ['FACTEUR DE PUISSANCE'],       'expand_rows': 1},
    {'pseudo_tag': 'TURBINE_VITESSE',        'label_contains': ['VITESSE TURBINE'],            'expand_rows': 1},
    {'pseudo_tag': 'TURBINE_SPEED_REF',      'label_contains': ['SPEED REF'],                  'expand_rows': 1},
    {'pseudo_tag': 'EXH_TEMP_MEAN',          'label_contains': ['MEAN'],                       'expand_rows': 1},
]
SPECIAL_PSEUDO_TAGS = {sf['pseudo_tag'] for sf in SPECIAL_FIELDS}


# ---------------------------------------------------------------------------
# 1) CONSTRUCTION DE LA CARTE DES CAPTEURS (tag -> cellule(s) Excel par heure)
# ---------------------------------------------------------------------------

def norm_tag(raw):
    """Extrait un tag canonique ('TE-532-AB') ou un pseudo-tag connu
    ('GEN_INTENSITE') depuis la reponse brute du modele."""
    if raw is None:
        return None
    s = str(raw).strip().upper()

    s_pseudo = re.sub(r'[\s\-]+', '_', s)
    if s_pseudo in SPECIAL_PSEUDO_TAGS:
        return s_pseudo

    s2 = re.sub(r'^\[[A-Z]\]\s*', '', s)      # retire prefixe [A]
    s2 = re.sub(r'^\d{2,4}-', '', s2)         # retire prefixe turbine ex "812-"
    return s2 if TAG_PATTERN.match(s2) else None


def find_label_anchor(ws, label_contains, label_excludes=None, restrict_rows=None):
    """Cherche une cellule texte contenant tous les mots de label_contains
    (et aucun de label_excludes). Retourne (row, col) ou None."""
    label_excludes = label_excludes or []
    for r in range(1, ws.max_row + 1):
        if restrict_rows is not None and r not in restrict_rows:
            continue
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if isinstance(v, str):
                vu = v.strip().upper()
                if all(sub in vu for sub in label_contains) and not any(sub in vu for sub in label_excludes):
                    return r, c
    return None


def build_sensor_map(xlsx_path):
    """Detecte automatiquement, pour chaque feuille du classeur, l'orientation
    (tags en colonnes / heures en lignes, ou l'inverse), construit :
       tag -> {sheet, cells: {heure: coordonnee_ou_liste_de_coordonnees}}
    pour les tags d'instrument classiques ET pour les champs speciaux
    sans tag (SPECIAL_FIELDS)."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    sensor_map = {}

    for sn in wb.sheetnames:
        ws = wb[sn]

        row_oriented_rows = {}   # heure -> numero de ligne (tags en colonnes, heures en lignes)
        for r in range(1, ws.max_row + 1):
            v = ws.cell(r, 1).value
            if isinstance(v, str) and v.strip() in HEURES:
                row_oriented_rows[v.strip()] = r

        col_oriented_cols = {}   # heure -> numero de colonne (tags en lignes, heures en colonnes)
        for r in range(1, min(ws.max_row, 6) + 1):
            for c in range(1, ws.max_column + 1):
                v = ws.cell(r, c).value
                if isinstance(v, str) and v.strip() in HEURES:
                    col_oriented_cols[v.strip()] = c

        if len(row_oriented_rows) >= 4:
            orientation = 'row'
        elif len(col_oriented_cols) >= 4:
            orientation = 'col'
        else:
            continue

        # --- tags d'instrument classiques ---
        if orientation == 'row':
            for r in range(1, ws.max_row + 1):
                for c in range(1, ws.max_column + 1):
                    tag = norm_tag(ws.cell(r, c).value)
                    if tag and tag not in sensor_map:
                        cells = {h: ws.cell(hr, c).coordinate for h, hr in row_oriented_rows.items()}
                        sensor_map[tag] = {'sheet': sn, 'cells': cells}
        else:
            for r in range(1, ws.max_row + 1):
                for c in range(1, ws.max_column + 1):
                    tag = norm_tag(ws.cell(r, c).value)
                    if tag and tag not in sensor_map:
                        cells = {h: ws.cell(r, hc).coordinate for h, hc in col_oriented_cols.items()}
                        sensor_map[tag] = {'sheet': sn, 'cells': cells}

        # --- champs speciaux sans tag d'instrument ---
        for sf in SPECIAL_FIELDS:
            pseudo = sf['pseudo_tag']
            if pseudo in sensor_map:
                continue
            if orientation == 'row':
                header_rows = set(range(1, min(row_oriented_rows.values())))
                anchor = find_label_anchor(ws, sf['label_contains'], sf.get('label_excludes'), restrict_rows=header_rows)
                if anchor:
                    _, c = anchor
                    cells = {h: ws.cell(hr, c).coordinate for h, hr in row_oriented_rows.items()}
                    sensor_map[pseudo] = {'sheet': sn, 'cells': cells}
            else:
                anchor = find_label_anchor(ws, sf['label_contains'], sf.get('label_excludes'))
                if anchor:
                    r, _ = anchor
                    expand = sf.get('expand_rows', 1)
                    if expand > 1:
                        cells = {h: [ws.cell(rr, hc).coordinate for rr in range(r, r + expand)]
                                 for h, hc in col_oriented_cols.items()}
                    else:
                        cells = {h: ws.cell(r, hc).coordinate for h, hc in col_oriented_cols.items()}
                    sensor_map[pseudo] = {'sheet': sn, 'cells': cells}

    return sensor_map


# ---------------------------------------------------------------------------
# 2) EXTRACTION DES VALEURS DEPUIS LES PHOTOS (Gemini vision)
# ---------------------------------------------------------------------------

def build_extraction_prompt(turbine=None, auto_detect_turbine=False):
    if auto_detect_turbine:
        turbine_clause = (
            'IMPORTANT - DETECTION DE LA TURBINE : ce lot de photos melange plusieurs '
            'turbines differentes. Pour CHAQUE ecran, lis le numero de turbine affiche '
            'dans le bandeau de titre (ex "GTG 812" -> turbine = "812") et indique-le '
            'dans le champ "turbine" du JSON (uniquement le numero, ex "812", pas "GTG 812").\n'
            'Cas particulier : si l\'ecran est un "Overview" qui affiche PLUSIEURS turbines '
            'en meme temps (ex plusieurs colonnes, une par turbine), laisse le champ '
            '"turbine" du haut a null, et indique a la place, pour CHAQUE valeur individuelle '
            'dans "readings", son propre champ "turbine" correspondant a la colonne dont elle '
            'provient. Si le numero de turbine n\'est pas clairement lisible pour une valeur, '
            'ne reporte PAS cette valeur plutot que de deviner sa turbine.'
        )
    else:
        turbine_clause = (
            f'Le numero de cette turbine est "{turbine}". Sur un ecran qui affiche plusieurs '
            f'turbines a la fois (ex ecran "Overview"), lis UNIQUEMENT la colonne/section de la '
            f'turbine {turbine}, ignore completement les autres turbines.'
            if turbine else
            'Le numero de turbine n\'a pas ete precise. Si un ecran affiche plusieurs turbines a '
            'la fois (ex ecran "Overview"), N\'EXTRAIS AUCUNE valeur de cet ecran (readings vide '
            'pour cette partie) plutot que de risquer de melanger les turbines.'
        )

    turbine_field_doc = (
        '  "turbine": "<numero de turbine, ex 812, ou null si non applicable/ambigu>",\n'
        if auto_detect_turbine else ''
    )
    reading_example = (
        '{"tag": "TE-532-AB", "value": 80.9, "unit": "C", "turbine": "812"}'
        if auto_detect_turbine else
        '{"tag": "TE-532-AB", "value": 80.9, "unit": "C"}'
    )

    return f"""Tu regardes une photo d'un ecran de supervision (HMI) d'une turbine a gaz.
{turbine_clause}

Ta tache : lire les paires (identifiant, valeur numerique) visibles a l'ecran.

=== CAS 1 : capteurs avec tag d'instrument (le cas le plus courant) ===
Les tags ressemblent a : TE-532-AB, PT-537, PDT-501, XT-610-X, KT-610, FT-511,
etc. (parfois prefixes par le numero de la turbine, ex "812-TE-532-AB" -> tag
a reporter = "TE-532-AB").

PATTERN FREQUENT A GERER : sur certains ecrans, un schema associe chaque
capteur a une lettre (A, B, C...) via une etiquette carree de couleur pres du
nom du capteur (ex carre vert "A" a cote de "812-TE-517-A"). La VALEUR de ce
capteur n'est pas forcement a cote du tag : elle est souvent ailleurs sur
l'ecran (graphique en barres, tableau...), reperee UNIQUEMENT par cette meme
lettre, sans que le tag y soit reecrit. Dans ce cas :
  1. Repere la legende qui associe chaque lettre a un tag complet.
  2. Repere ailleurs sur l'ecran la valeur numerique sous/pres de cette lettre.
  3. Associe les deux -> (tag complet, valeur).
Ignore les colonnes "SEL" (valeur selectionnee/calculee, pas un capteur).
En cas de doute sur la correspondance lettre<->tag, ne reporte pas la paire.

=== CAS 2 : champs SANS tag d'instrument, juste un nom de parametre ===
Utilise EXACTEMENT le pseudo-tag indique comme "tag" dans le JSON :

  Ecran "Overview" (parametres electriques generateur) :
    Intensite / courant generateur (A)   -> "GEN_INTENSITE"
    Tension generateur (kV)              -> "GEN_TENSION"
    Frequence (Hz)                       -> "GEN_FREQUENCE"
    Puissance active (MW/MV)             -> "GEN_PUISSANCE_ACTIVE"
    Puissance reactive (MVAR)            -> "GEN_PUISSANCE_REACTIVE"
    Facteur de puissance / cos phi       -> "GEN_FACTEUR_PUISSANCE"

  Ecran "Start-up" :
    Vitesse turbine (rpm)                -> "TURBINE_VITESSE"
    Speed ref / vitesse de reference     -> "TURBINE_SPEED_REF"

  Ecran "Turbine Exhaust Temp" :
    Valeur "Mean" / temperature moyenne d'echappement -> "EXH_TEMP_MEAN"

Si un ecran ne correspond a aucun de ces cas particuliers, utilise le tag
d'instrument normal du CAS 1.

Regles strictes :
- N'invente et ne devine JAMAIS une valeur. Si un chiffre est flou, coupe ou
  ambigu, ne le reporte PAS.
- Ignore les elements qui ne sont pas des valeurs numeriques de capteur
  (boutons, statuts "NORMAL"/"REMOTE", noms d'ecran, alarmes texte...).
- Un capteur = un identifiant (tag ou pseudo-tag) + une valeur numerique
  (garde l'unite a part).

Reponds UNIQUEMENT avec un JSON valide (rien d'autre, pas de balises ```), au format :
{{
  "screen_title": "<titre de l'ecran tel qu'affiche>",
{turbine_field_doc}  "readings": [
    {reading_example},
    {{"tag": "GEN_INTENSITE", "value": 1250, "unit": "A"}}
  ]
}}

Si aucune valeur fiable n'est lisible, renvoie une liste "readings" vide.
"""


def extract_from_image_gemini(client, image_path, model_name, prompt, max_retries=4):
    """Envoie une image a Gemini et recupere la liste des (tag, valeur) lus."""
    ext = Path(image_path).suffix.lower()
    mime_type = {
        '.jpg': 'image/jpeg', '.jpeg': 'image/jpeg',
        '.png': 'image/png', '.webp': 'image/webp',
    }.get(ext, 'image/jpeg')

    with open(image_path, 'rb') as f:
        img_bytes = f.read()

    last_err = None
    quota_error = False
    for attempt in range(1, max_retries + 1):
        try:
            response = client.models.generate_content(
                model=model_name,
                contents=[
                    genai_types.Part.from_bytes(data=img_bytes, mime_type=mime_type),
                    prompt,
                ],
            )
            text = (response.text or "").strip()
            text = re.sub(r'^```(json)?|```$', '', text, flags=re.MULTILINE).strip()
            return json.loads(text)
        except json.JSONDecodeError:
            print(f"  [ATTENTION] Reponse non-JSON pour {image_path}, ignoree.")
            return {"screen_title": None, "readings": []}
        except Exception as e:
            last_err = e
            quota_error = False
            msg = str(e)
            if '429' in msg or 'RESOURCE_EXHAUSTED' in msg or 'rate' in msg.lower():
                quota_error = True
                wait = 20 * attempt
                print(f"  [Limite gratuite atteinte] Nouvelle tentative dans {wait}s "
                      f"({attempt}/{max_retries}) ...")
                time.sleep(wait)
                continue
            if '503' in msg or 'UNAVAILABLE' in msg or 'overloaded' in msg.lower() or 'high demand' in msg.lower():
                wait = 15 * attempt
                print(f"  [Serveur Gemini surcharge] Nouvelle tentative dans {wait}s "
                      f"({attempt}/{max_retries}) ...")
                time.sleep(wait)
                continue
            break

    if quota_error:
        raise QuotaExhaustedError(str(last_err))

    print(f"  [ERREUR] Echec de lecture de {image_path} : {last_err}")
    return {"screen_title": None, "readings": []}


def extract_from_image_ollama(image_path, model_name, prompt, ollama_url="http://localhost:11434", max_retries=2):
    """Envoie une image a un modele vision local via Ollama (100% hors-ligne, gratuit,
    sans limite). Necessite que Ollama tourne en arriere-plan (ollama.com) et que le
    modele ait deja ete telecharge (ex: ollama pull qwen2.5vl:3b)."""
    try:
        import requests
    except ImportError:
        sys.exit("Le package 'requests' n'est pas installe. Faites : pip install requests")

    with open(image_path, 'rb') as f:
        img_b64 = base64.b64encode(f.read()).decode('utf-8')

    last_err = None
    for attempt in range(1, max_retries + 1):
        try:
            resp = requests.post(
                f"{ollama_url}/api/chat",
                json={
                    "model": model_name,
                    "messages": [{"role": "user", "content": prompt, "images": [img_b64]}],
                    "stream": False,
                },
                timeout=300,  # l'inference CPU locale peut etre lente, on laisse le temps
            )
            resp.raise_for_status()
            text = resp.json().get("message", {}).get("content", "").strip()
            text = re.sub(r'^```(json)?|```$', '', text, flags=re.MULTILINE).strip()
            return json.loads(text)
        except json.JSONDecodeError:
            print(f"  [ATTENTION] Reponse non-JSON pour {image_path}, ignoree.")
            return {"screen_title": None, "readings": []}
        except requests.exceptions.ConnectionError:
            sys.exit(
                "Impossible de contacter Ollama sur " + ollama_url + ".\n"
                "Verifiez qu'Ollama est bien installe et lance (il tourne normalement en\n"
                "arriere-plan automatiquement apres l'installation). Testez avec :\n"
                "    ollama run " + model_name
            )
        except Exception as e:
            last_err = e
            print(f"  [Nouvelle tentative apres erreur locale] {e}")
            time.sleep(5)

    print(f"  [ERREUR] Echec de lecture de {image_path} : {last_err}")
    return {"screen_title": None, "readings": []}


def extract_from_image_mistral(image_path, model_name, prompt, api_key, max_retries=3):
    """Envoie une image a l'API vision gratuite de Mistral (Pixtral). Fournisseur
    totalement independant de Google : utile en secours quand Gemini est sature."""
    try:
        import requests
    except ImportError:
        sys.exit("Le package 'requests' n'est pas installe. Faites : pip install requests")

    ext = Path(image_path).suffix.lower()
    mime_type = {
        '.jpg': 'image/jpeg', '.jpeg': 'image/jpeg',
        '.png': 'image/png', '.webp': 'image/webp',
    }.get(ext, 'image/jpeg')

    with open(image_path, 'rb') as f:
        img_b64 = base64.b64encode(f.read()).decode('utf-8')

    last_err = None
    quota_error = False
    for attempt in range(1, max_retries + 1):
        try:
            resp = requests.post(
                "https://api.mistral.ai/v1/chat/completions",
                headers={"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"},
                json={
                    "model": model_name,
                    "messages": [{
                        "role": "user",
                        "content": [
                            {"type": "text", "text": prompt},
                            {"type": "image_url", "image_url": {"url": f"data:{mime_type};base64,{img_b64}"}},
                        ],
                    }],
                },
                timeout=90,
            )
            if resp.status_code == 429:
                raise RuntimeError(f"429 rate limit / quota: {resp.text[:200]}")
            if resp.status_code >= 500:
                raise RuntimeError(f"{resp.status_code} serveur Mistral indisponible: {resp.text[:200]}")
            resp.raise_for_status()
            text = resp.json()["choices"][0]["message"]["content"].strip()
            text = re.sub(r'^```(json)?|```$', '', text, flags=re.MULTILINE).strip()
            return json.loads(text)
        except json.JSONDecodeError:
            print(f"  [ATTENTION] Reponse non-JSON pour {image_path}, ignoree.")
            return {"screen_title": None, "readings": []}
        except Exception as e:
            last_err = e
            quota_error = False
            msg = str(e)
            if '429' in msg or 'rate' in msg.lower() or 'quota' in msg.lower():
                quota_error = True
                wait = 20 * attempt
                print(f"  [Limite Mistral atteinte] Nouvelle tentative dans {wait}s ({attempt}/{max_retries}) ...")
                time.sleep(wait)
                continue
            if 'indisponible' in msg.lower() or 'timeout' in msg.lower():
                wait = 15 * attempt
                print(f"  [Serveur Mistral surcharge] Nouvelle tentative dans {wait}s ({attempt}/{max_retries}) ...")
                time.sleep(wait)
                continue
            break

    if quota_error:
        raise QuotaExhaustedError(str(last_err))

    print(f"  [ERREUR] Echec de lecture de {image_path} : {last_err}")
    return {"screen_title": None, "readings": []}

def fill_workbook(xlsx_path, sensor_map, heure, image_paths, out_path, api_key=None,
                   model_name="gemini-3.5-flash", turbine=None, resume=False,
                   backend="gemini", ollama_url="http://localhost:11434",
                   fallback_backend=None, fallback_api_key=None, fallback_model=None):

    def make_client(be, key):
        if be == "gemini":
            if genai is None:
                sys.exit("Le package 'google-genai' n'est pas installe. Faites : pip install google-genai")
            key = key or os.environ.get("GEMINI_API_KEY")
            if not key:
                sys.exit("Aucune cle API Gemini. Definissez GEMINI_API_KEY ou utilisez --api-key.\n"
                          "Cle gratuite ici : https://aistudio.google.com/apikey")
            return genai.Client(api_key=key)
        elif be == "mistral":
            key = key or os.environ.get("MISTRAL_API_KEY")
            if not key:
                sys.exit("Aucune cle API Mistral. Definissez MISTRAL_API_KEY ou utilisez --fallback-api-key.\n"
                          "Cle gratuite ici : https://console.mistral.ai/api-keys")
            return key  # pour mistral, le "client" est juste la cle (appels via requests)
        return None  # ollama n'a pas besoin de client

    def default_model(be):
        return {"gemini": "gemini-3.5-flash", "mistral": "pixtral-large-latest", "ollama": "qwen2.5vl:3b"}[be]

    def run_extraction(be, client_or_key, mdl, img_path, prompt):
        if be == "gemini":
            r = extract_from_image_gemini(client_or_key, img_path, mdl, prompt)
            time.sleep(3)  # reste dans les limites gratuites (~10 requetes/minute)
            return r
        elif be == "mistral":
            return extract_from_image_mistral(img_path, mdl, prompt, client_or_key)
        else:
            return extract_from_image_ollama(img_path, mdl, prompt, ollama_url)

    active_backend = backend
    active_client = make_client(backend, api_key)
    active_model = model_name or default_model(backend)

    fb_client = None
    if fallback_backend and fallback_backend != "none":
        fb_client = make_client(fallback_backend, fallback_api_key)
        fb_model = fallback_model or default_model(fallback_backend)
        print(f"[INFO] Secours automatique active : si '{active_backend}' est sature, "
              f"le script basculera tout seul sur '{fallback_backend}' pour la suite.\n")

    prompt = build_extraction_prompt(turbine)
    progress_path = out_path + ".progress.json"

    # --- reprise apres arret (nouvelle cle API, panne, etc.) ---
    # IMPORTANT : on ne recharge JAMAIS out_path (openpyxl ne supporte pas de
    # re-sauvegarder plusieurs fois un classeur contenant une image embarquee
    # dans le meme run -> on repart toujours du modele original et on rejoue
    # les valeurs deja trouvees, puis on ne sauvegarde le .xlsx qu'UNE SEULE
    # FOIS a la fin (ou au moment de l'arret pour quota epuise).
    processed = set()
    matched, unmatched_from_photos, seen_tags = [], [], set()
    if resume and os.path.exists(progress_path):
        with open(progress_path, encoding='utf-8') as f:
            state = json.load(f)
        processed = set(state.get("processed", []))
        matched = [tuple(m) for m in state.get("matched", [])]
        unmatched_from_photos = [tuple(m) for m in state.get("unmatched", [])]
        seen_tags = set(state.get("seen_tags", []))
        print(f"[REPRISE] {len(processed)} photo(s) deja traitee(s), on continue avec les {len(image_paths) - len(processed)} restante(s).\n")

    wb = openpyxl.load_workbook(xlsx_path)  # garde les formules/formatage
    for tag, value, sheet_name, coord_display, _title in matched:
        ws = wb[sheet_name]
        for c in coord_display.split('+'):
            ws[c] = value

    def save_progress_json():
        # Sauvegarde legere (texte seulement) apres chaque photo : rien n'est
        # perdu en cas d'arret/crash, sans toucher au fichier .xlsx.
        with open(progress_path, 'w', encoding='utf-8') as f:
            json.dump({
                "processed": sorted(processed),
                "matched": matched,
                "unmatched": unmatched_from_photos,
                "seen_tags": sorted(seen_tags),
            }, f, ensure_ascii=False)

    for i, img_path in enumerate(image_paths, 1):
        img_key = str(img_path)
        if img_key in processed:
            print(f"[{i}/{len(image_paths)}] {Path(img_path).name} deja traitee, on saute.")
            continue

        print(f"[{i}/{len(image_paths)}] Lecture de {Path(img_path).name} (via {active_backend}) ...")
        try:
            result = run_extraction(active_backend, active_client, active_model, img_path, prompt)
        except QuotaExhaustedError as qe:
            if fb_client is not None and active_backend != fallback_backend:
                print(f"\n[BASCULEMENT] '{active_backend}' semble sature ({qe}).")
                print(f"[BASCULEMENT] Passage automatique sur '{fallback_backend}' pour la suite de la tournee...\n")
                active_backend = fallback_backend
                active_client = fb_client
                active_model = fb_model
                try:
                    result = run_extraction(active_backend, active_client, active_model, img_path, prompt)
                except QuotaExhaustedError as qe2:
                    wb.save(out_path)
                    save_progress_json()
                    print("\n" + "=" * 70)
                    print(f"[QUOTA EPUISE] Les DEUX backends ('{backend}' et '{fallback_backend}') semblent satures :")
                    print(f"   {qe2}")
                    print(f"\nProgres sauvegarde : {len(processed)}/{len(image_paths)} photo(s) deja traitees.")
                    print(f"Fichier partiel enregistre : {out_path}")
                    print("\nMARCHE A SUIVRE : changez de cle(s) API puis relancez avec --resume.")
                    print("=" * 70)
                    sys.exit(1)
            else:
                wb.save(out_path)
                save_progress_json()
                print("\n" + "=" * 70)
                print(f"[QUOTA EPUISE] Le backend '{active_backend}' semble reellement sature")
                print("(pas juste une limite temporaire) :")
                print(f"   {qe}")
                print(f"\nProgres sauvegarde : {len(processed)}/{len(image_paths)} photo(s) deja traitees.")
                print(f"Fichier partiel enregistre : {out_path}")
                print("\nMARCHE A SUIVRE :")
                print("  1. Changez de cle API (nouvelle cle gratuite), ou utilisez --fallback-backend "
                      "pour basculer automatiquement sur un 2e fournisseur la prochaine fois.")
                print("  2. Relancez EXACTEMENT la meme commande en ajoutant --resume et la nouvelle --api-key")
                print("  3. Le script reprendra automatiquement a la photo suivante, sans repeter le travail deja fait.")
                print("=" * 70)
                sys.exit(1)

        title = result.get("screen_title") or "(titre non lu)"
        readings = result.get("readings", [])
        print(f"    Ecran: {title} -> {len(readings)} valeur(s) lue(s)")

        for reading in readings:
            tag = norm_tag(reading.get("tag"))
            value = reading.get("value")
            if tag is None or value is None:
                continue
            seen_tags.add(tag)

            if tag in sensor_map:
                sheet_name = sensor_map[tag]['sheet']
                coord = sensor_map[tag]['cells'].get(heure)
                if coord is None:
                    unmatched_from_photos.append((tag, value, title, "heure introuvable dans la carte"))
                    continue
                ws = wb[sheet_name]
                if isinstance(coord, list):
                    for c in coord:
                        ws[c] = value
                    coord_display = "+".join(coord)
                else:
                    ws[coord] = value
                    coord_display = coord
                matched.append((tag, value, sheet_name, coord_display, title))
            else:
                unmatched_from_photos.append((tag, value, title, "tag absent du classeur Excel"))

        processed.add(img_key)
        save_progress_json()  # sauvegarde legere apres CHAQUE photo (pas de re-sauvegarde du .xlsx)

    wb.save(out_path)  # UNE SEULE sauvegarde du .xlsx pour toute la tournee
    if os.path.exists(progress_path):
        os.remove(progress_path)  # tournee terminee avec succes, plus besoin du fichier de reprise

    # --- rapport ---
    print("\n" + "=" * 70)
    print(f"TERMINE. Fichier enregistre : {out_path}")
    print("=" * 70)
    print(f"\n[OK] {len(matched)} valeur(s) ecrite(s) avec succes :")
    for tag, value, sheet, coord, title in matched:
        print(f"   {tag:22s} = {value:<10} -> {sheet}!{coord}   (depuis: {title})")

    if unmatched_from_photos:
        print(f"\n[A VERIFIER] {len(unmatched_from_photos)} valeur(s) lue(s) sur les photos "
              f"mais NON ecrites :")
        for tag, value, title, reason in unmatched_from_photos:
            print(f"   {tag:22s} = {value:<10}   ({reason}, depuis: {title})")

    expected_not_seen = sorted(set(sensor_map.keys()) - seen_tags)
    if expected_not_seen:
        print(f"\n[MANQUANT] {len(expected_not_seen)} capteur(s)/champ(s) attendu(s) dans le "
              f"classeur mais jamais vus sur une photo :")
        for tag in expected_not_seen:
            print(f"   {tag}")

    print("\nRien n'a ete devine : verifiez les sections ci-dessus avant de valider le rapport.")


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description="Remplit un rapport GTG Excel a partir de photos d'ecran (via Gemini vision).")
    sub = parser.add_subparsers(dest='command', required=True)

    p_map = sub.add_parser('map', help="Genere la carte des capteurs a partir du fichier Excel")
    p_map.add_argument('--excel', required=True)
    p_map.add_argument('--out', default='sensor_map.json')

    p_fill = sub.add_parser('fill', help="Remplit le classeur a partir de photos")
    p_fill.add_argument('--excel', required=True, help="Fichier Excel modele/a remplir")
    p_fill.add_argument('--map', required=True, help="Fichier sensor_map.json (genere par la commande 'map')")
    p_fill.add_argument('--heure', required=True, choices=HEURES, help="Heure du releve : 4h, 8h, 12h, 16h, 20h ou 24h")
    p_fill.add_argument('--turbine', default=None, help="Numero de la turbine (ex 812), pour lire la bonne colonne sur l'ecran Overview multi-turbines")
    p_fill.add_argument('--images', nargs='*', default=[], help="Liste de chemins vers les photos")
    p_fill.add_argument('--images-dir', default=None, help="Dossier contenant les photos (alternative a --images)")
    p_fill.add_argument('--out', required=True, help="Fichier Excel de sortie")
    p_fill.add_argument('--api-key', default=None, help="Cle API du backend principal (sinon variable d'env GEMINI_API_KEY/MISTRAL_API_KEY selon --backend)")
    p_fill.add_argument('--model', default=None, help="Modele a utiliser pour le backend principal (defaut selon --backend)")
    p_fill.add_argument('--backend', choices=['gemini', 'mistral', 'ollama'], default='gemini',
                         help="'gemini' = API gratuite Google (defaut). 'mistral' = API gratuite Mistral/Pixtral. "
                              "'ollama' = modele local sur votre PC, sans limite, necessite Ollama installe.")
    p_fill.add_argument('--ollama-url', default='http://localhost:11434', help="Adresse du serveur Ollama local (defaut: http://localhost:11434)")
    p_fill.add_argument('--fallback-backend', choices=['none', 'gemini', 'mistral', 'ollama'], default='none',
                         help="Backend de secours : si le backend principal (--backend) est sature/en quota "
                              "epuise, le script bascule AUTOMATIQUEMENT dessus pour le reste de la tournee, "
                              "sans s'arreter. Ex: --backend gemini --fallback-backend mistral")
    p_fill.add_argument('--fallback-api-key', default=None, help="Cle API du backend de secours (sinon variable d'env correspondante)")
    p_fill.add_argument('--fallback-model', default=None, help="Modele a utiliser pour le backend de secours (defaut selon --fallback-backend)")
    p_fill.add_argument('--resume', action='store_true',
                         help="Reprend une tournee interrompue (ex: quota epuise) la ou elle s'est arretee, "
                              "en reutilisant le meme --out. Combinez avec une nouvelle --api-key si besoin.")

    args = parser.parse_args()

    if args.command == 'map':
        sm = build_sensor_map(args.excel)
        with open(args.out, 'w', encoding='utf-8') as f:
            json.dump(sm, f, indent=2, ensure_ascii=False)
        print(f"{len(sm)} capteur(s)/champ(s) detecte(s) -> {args.out}")

    elif args.command == 'fill':
        with open(args.map, encoding='utf-8') as f:
            sensor_map = json.load(f)

        image_paths = list(args.images)
        if args.images_dir:
            exts = {'.jpg', '.jpeg', '.png', '.webp'}
            image_paths += sorted(str(p) for p in Path(args.images_dir).iterdir() if p.suffix.lower() in exts)

        if not image_paths:
            sys.exit("Aucune image fournie (utilisez --images ou --images-dir).")

        fallback_backend = None if args.fallback_backend == 'none' else args.fallback_backend

        fill_workbook(args.excel, sensor_map, args.heure, image_paths, args.out,
                      api_key=args.api_key, model_name=args.model, turbine=args.turbine,
                      resume=args.resume, backend=args.backend, ollama_url=args.ollama_url,
                      fallback_backend=fallback_backend, fallback_api_key=args.fallback_api_key,
                      fallback_model=args.fallback_model)


if __name__ == '__main__':
    main()
